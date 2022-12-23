import pyrebase
firebaseConfig = {
  "apiKey": "AIzaSyD0KS2N_37H5Pikp3pkAS5BhLkzVqMMIfE",
  "authDomain": "fir-demo-9f71c.firebaseapp.com",
  "databaseURL": "https://fir-demo-9f71c-default-rtdb.firebaseio.com/",
  "projectId": "fir-demo-9f71c",
  "storageBucket": "fir-demo-9f71c.appspot.com",
  "messagingSenderId": "586782017358",
  "appId": "1:586782017358:web:2c324717d215942025001f",
  "measurementId": "G-DRSD92QYY1"
}
firebase=pyrebase.initialize_app(firebaseConfig)
db=firebase.database()

from flask import Flask,render_template,request,redirect,flash
app = Flask(__name__)
import openpyxl

final_menu={} #key->[item_name,visitor_price,staff_price]

helper_dictionary={} #item_name-> key

#Excel function
def addItemsFromExcelToDb():
    path='menu.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    row = sheet_obj.max_row
    col = sheet_obj.max_column

    for i in range(4,row+1):
        item_obj = sheet_obj.cell(row=i, column=2)
        vistor_obj = sheet_obj.cell(row=i, column=3)
        staff_obj = sheet_obj.cell(row=i, column=4)

        #print(i," ",item_obj.value,vistor_obj.value,staff_obj.value)

        item_str=str(item_obj.value)
        #print(item_str)
        # item_str=item_str.replace(' ','-')
        # item_str=item_str.replace('/','_')

        helper_dictionary[item_str]=i
        final_menu[i]=(item_str,vistor_obj.value,staff_obj.value)

    print(final_menu)
    db.child('final_menu').set(final_menu)

def getUid(item):
    print(helper_dictionary)
    if item in helper_dictionary:
        return helper_dictionary[item]
    else:
        return None

#excel function
def getItemsFromMenu():
    path= "menu.xlsx"

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active


    row=sheet_obj.max_row
    col=sheet_obj.max_column


    for i in range(4, row + 1):
        item_obj = sheet_obj.cell(row = i, column = 2)
        vistor_obj=sheet_obj.cell(row=i,column=3)
        staff_obj=sheet_obj.cell(row=i,column=4)

        item_str=str(item_obj.value)
        print(item_str)
        final_menu[item_str]=i
       # final_menu_db[i]=[item_str,vistor_obj.value,staff_obj.value]

#Initializes final_menu : uid->[item_name,staff_price,visitor_price]  -->Stored in DB
# helper_dictionary: item_name->uid --> Stored in Program for fast access
def initialization():
    global final_menu
    global helper_dictionary
    final_menu_obj = db.child('final_menu').get()
    for menu_item_details in final_menu_obj.each():
        items_lst=db.child('final_menu').child(menu_item_details.key()).get().val()
        uid=menu_item_details.key()
        if items_lst is not None:
            final_menu[uid]=items_lst
            helper_dictionary[items_lst[0]]=uid



def getMenuItemsFromCurrentMenu():
    global final_menu
    final_menu_lst=[]
    for k,v in final_menu.items():
        final_menu_lst.append([k]+v)
    clean_lst=[ele for ele in final_menu_lst if ele is not None]

    return clean_lst


    #db.update({"final_menu":final_menu_db})

#making this python file like a API
#addItemsFromExcelToDb()

@app.route("/menu/<uid>/<prize>/add")
def addToCart(uid,prize):
    print(uid,prize)
    return redirect("/menu")

@app.route("/menu/harshit")
def something():
    print("agalg")

@app.route("/menu")
def menuMain():
    menu_items=getMenuItemsFromCurrentMenu()
    return render_template("menu_index.html",menu_items=menu_items)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    initialization()
    app.run(debug=True)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
